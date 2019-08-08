import openpyxl

#load workbook

wk = openpyxl.load_workbook('D:\\Book1.xlsx', 'r')
print(wk.sheetnames)
print("Active Sheet is " + wk.active.title)

#Create object of any sheet on which you want to work

sh = wk['Sheet1']
print(sh.title)

#Fetch the value
print(sh['A4'].value)
print(sh['B3'].value)
print(sh['C2'].value)

#creating cell object
cl = sh.cell(4,2)
print(cl.value)

#Find Rows having data
rows = sh.max_row
columns = sh.max_column

print("Total Rows are - " + str(rows))
print("Total Columns are - " + str(columns))

for i in range(1,rows+1):
    for j in range(1,columns+1):
     c = sh.cell(i,j)
    print(c.value)

    #for r in sh['A1' : 'C4']:
    #for r in sh['A1' : 'C4']:
    #for c in r:
    #print(c.value)








